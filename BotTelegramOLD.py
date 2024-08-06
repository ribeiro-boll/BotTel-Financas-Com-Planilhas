import telebot
from telebot import types
from keys import keytel
import integracaoExcel
from integracaoExcel import seuMes
import datetime
bot = telebot.TeleBot(keytel())
data = str(datetime.datetime.now())
pessoa = ''
meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
valorNome = ''
valorGasto = 0
valorExcel = []
valorTotal = 0
mesSelecionado = []
valorNomeEditar = []
valorNomeEditar2 = []
opcaoNumero = []
arquivo = []


@bot.message_handler(commands=["selecionarPlanilha"])
def criarplanilha(message):
    from integracaoExcel import opcoesAno
    arquivo.clear()
    chat_id = message.chat.id
    bot.send_message(message.chat.id, "Você pode escolher entre: \n\n 7 anos atrás ou 3 anos no futuro.")
    a12 = opcoesAno()
    anosEscolha = types.ReplyKeyboardMarkup()
    c = types.KeyboardButton(f'{a12[0]}')
    d = types.KeyboardButton(f'{a12[1]}')
    e = types.KeyboardButton(f'{a12[2]}')
    f = types.KeyboardButton(f'{a12[3]}')
    g = types.KeyboardButton(f'{a12[4]}')
    h = types.KeyboardButton(f'{a12[5]}')
    i = types.KeyboardButton(f'{a12[6]}')
    j = types.KeyboardButton(f'{a12[7]}')
    k = types.KeyboardButton(f'{a12[8]}')
    l = types.KeyboardButton(f'{a12[9]}')
    m = types.KeyboardButton(f'{a12[10]}')
    anosEscolha.row(h)
    anosEscolha.row(c, d, e, f, g)
    anosEscolha.row(i, j, k, l, m)
    anoParte1 = bot.send_message(chat_id, "selecione o ano \n\n (obs: a primeira e maior opção, sera o ano atual)", reply_markup=anosEscolha)
    bot.register_next_step_handler(anoParte1, anoPlanilha1)


def anoPlanilha1(message):
    if int(message.text) > 2010 and int(message.text) < 2099 and len(message.text) == 4:
        pessoa = message.from_user.first_name
        chat_id = message.chat.id
        ano = message.text
        a = integracaoExcel.criarArquivoAno(pessoa, ano)
        arquivo.append(a)
        bot.send_message(chat_id, "Sua planilha foi criada/acessada com sucesso! \n\n Para começar a usar o programa, aperte em \n-> /comecar")
    else:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/comecar')
        b = types.KeyboardButton('/selecionarPlanilha')
        voltar.row(a)
        voltar.row(b)
        bot.send_message(chat_id, "Algo deu errado \n\n Para voltar para o início, aperte \n-> /comecar \n\n Para tentar novamente, aperte \n-> /selecionarPlanilha", reply_markup=voltar)

@bot.message_handler(commands=["comecar"])
def greetings(message):
    pessoa = message.from_user.first_name
    import datetime
    data = str(datetime.datetime.now())
    chat_id = message.chat.id
    if arquivo == []:
        planilha = integracaoExcel.criarArquivoAno(pessoa, data[0] + data[1] + data[2] + data[3])
        arquivo.append(planilha)
        listadoida = list(arquivo[0])
        opcao = types.ReplyKeyboardMarkup()
        ano = listadoida[-9:-5]
        ano = ''.join(ano)
        a = types.KeyboardButton('/addGasto')
        b = types.KeyboardButton('/contasMes')
        c = types.KeyboardButton('/editarVer')
        d = types.KeyboardButton('/selecionarPlanilha')
        opcao.row(d)
        opcao.row(a, b, c)
        bot.send_message(chat_id, "O limite de gastos por mês é de 50 gastos.")
        bot.send_message(chat_id,
                         f"Selecione sua opção: \n\n Para adicionar gasto em sua planilha \n-> /addGasto \n\n Caso queira ver o total de gastos no mês, toque em \n-> /contasMes \n\n Caso queira editar ou ver seus gastos, toque em \n-> /editarVer \n\n Sua planilha atual é do ano \n\n Para mudar a planilha, aperte em \n-> /selecionarPlanilha \n\n Sua planilha atual é do ano: {ano}", reply_markup=opcao)
        listadoida.clear()
    else:
        listadoida = list(arquivo[0])
        ano = listadoida[-9:-5]
        ano = ''.join(ano)
        opcao = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/addGasto')
        b = types.KeyboardButton('/contasMes')
        c = types.KeyboardButton('/editarVer')
        d = types.KeyboardButton('/selecionarPlanilha')
        opcao.row(d)
        opcao.row(a, b, c)
        bot.send_message(chat_id, "O limite de gastos por mês é de 50 gastos.")
        bot.send_message(chat_id,
                         f"Selecione sua opção: \n\n Para adicionar gasto em sua planilha \n-> /addGasto \n\n Caso queira ver o total de gastos no mês, toque em \n-> /contasMes \n\n Caso queira editar ou ver seus gastos, toque em \n-> /editarVer \n\n Sua planilha atual é do ano \n\n Para mudar a planilha, aperte em \n-> /selecionarPlanilha \n\n Sua planilha atual é do ano:{ano}",
                         reply_markup=opcao)
        listadoida.clear()

@bot.message_handler(commands=["editarVer"])
def editarVer(message):
    chat_id = message.chat.id
    meses = types.ReplyKeyboardMarkup()
    a = types.KeyboardButton('jan'.capitalize())
    b = types.KeyboardButton('fev'.capitalize())
    c = types.KeyboardButton('mar'.capitalize())
    d = types.KeyboardButton('abr'.capitalize())
    e = types.KeyboardButton('mai'.capitalize())
    f = types.KeyboardButton('jun'.capitalize())
    g = types.KeyboardButton('jul'.capitalize())
    h = types.KeyboardButton('ago'.capitalize())
    i = types.KeyboardButton('set'.capitalize())
    j = types.KeyboardButton('out'.capitalize())
    k = types.KeyboardButton('nov'.capitalize())
    l = types.KeyboardButton('dez'.capitalize())
    m = types.KeyboardButton(seuMes().capitalize())
    meses.row(m)
    meses.row(a, b, c, d)
    meses.row(e, f, g, h)
    meses.row(i, j, k, l)
    mes = bot.send_message(chat_id,"Selecione o mês \n\n (obs: a primeira e maior opção será o mês atual)", reply_markup=meses)
    bot.register_next_step_handler(mes, editar1)

def editar1(message):
   try:
       chat_id = message.chat.id
       from openpyxl import load_workbook
       from openpyxl.utils import get_column_letter
       import datetime
       wb = load_workbook(arquivo[0])
       ws = wb[message.text]
       coluna = get_column_letter(2)
       colunaNome = get_column_letter(1)
       colunaTexto = ''
       for row in range(1, 50):
           valor = ws[colunaNome + str(row)].value
           if valor == None:
               continue
           else:
               colunaTexto += f'{ws[colunaNome + str(row)].value} = R${float(ws[coluna + str(row)].value)}\n\n'
       bot.send_message(chat_id, colunaTexto)
       voltar = types.ReplyKeyboardMarkup()
       a = types.KeyboardButton('/comecar')
       b = types.KeyboardButton("/editar")
       voltar.row(a, b)
       bot.send_message(chat_id, "Para editar algum valor, aperte em \n-> /editar \n\n Para voltar ao início, aperte \n-> /comecar", reply_markup=voltar)
   except KeyError:
       chat_id = message.chat.id
       voltar = types.ReplyKeyboardMarkup()
       a = types.KeyboardButton('/comecar')
       b = types.KeyboardButton("/editarVer")
       voltar.row(a, b)
       bot.send_message(chat_id, "Algo deu errado \n\n Para voltar para o início, aperte \n-> /comecar \n\n Para tentar novamente, aperte \n-> /editarVer", reply_markup=voltar)



@bot.message_handler(commands=["editar"])
def editarFunc1(message):
    mesSelecionado.clear()
    valorNomeEditar.clear()
    valorNomeEditar2.clear()
    opcaoNumero.clear()
    chat_id = message.chat.id
    meses = types.ReplyKeyboardMarkup()
    a = types.KeyboardButton('jan'.capitalize())
    b = types.KeyboardButton('fev'.capitalize())
    c = types.KeyboardButton('mar'.capitalize())
    d = types.KeyboardButton('abr'.capitalize())
    e = types.KeyboardButton('mai'.capitalize())
    f = types.KeyboardButton('jun'.capitalize())
    g = types.KeyboardButton('jul'.capitalize())
    h = types.KeyboardButton('ago'.capitalize())
    i = types.KeyboardButton('set'.capitalize())
    j = types.KeyboardButton('out'.capitalize())
    k = types.KeyboardButton('nov'.capitalize())
    l = types.KeyboardButton('dez'.capitalize())
    m = types.KeyboardButton(seuMes().capitalize())
    meses.row(m)
    meses.row(a, b, c, d)
    meses.row(e, f, g, h)
    meses.row(i, j, k, l)
    mes = bot.send_message(chat_id,"Selecione o mês \n\n (obs: a primeira e maior opção será o mês atual)", reply_markup=meses)
    bot.register_next_step_handler(mes, editarFunc2)

def editarFunc2(message):
    try:
        mesSelecionado.append(str(message.text))
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        chat_id = message.chat.id
        a = get_column_letter(1)
        wb = load_workbook(arquivo[0])
        ws = wb[message.text]
        nomes = ''
        markup = types.ForceReply()
        for row in range(1, 50):
            valor = ws[a + str(row)].value
            if valor == None:
                continue
            else:
                nomes += f'{ws[a + str(row)].value} gasto nº /0000{row}\n\n'
        nomesGastosEditar = bot.send_message(chat_id,
                                             f'{nomes}\ntoque no gasto que sera editado ou digite a sua numeração com uma barra na frente\n\nEx: Exemplo gasto nº /00001 - equivale ao gasto nº1',
                                             reply_markup=markup)
        bot.register_next_step_handler(nomesGastosEditar, editarFunc3)
    except KeyError:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/editar')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,
                         "Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n\n Tente novamente com \n-> /editar \n\n Caso queira voltar para o início, toque em \n-> /comecar",
                         reply_markup=voltar)


def editarFunc3(message):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        chat_id = message.chat.id
        var = list(message.text)
        var.remove("/")
        var.remove("0")
        var.remove("0")
        var.remove("0")
        var.remove("0")
        a = ''.join(var)
        print(a)
        numeroEscolhido = str(a)
        valorNomeEditar.append(numeroEscolhido)
        wb = load_workbook(arquivo[0])
        ws = wb[mesSelecionado[0]]
        a = get_column_letter(1)
        nnome = ws[a + str(valorNomeEditar[0])].value
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('Nome')
        b = types.KeyboardButton('Valor')
        voltar.row(a, b)
        opcao = bot.send_message(chat_id, f"Para editar {nnome}, aperte em \n-> Nome. Para editar o valor, aperte em \n-> Valor",
                                 reply_markup=voltar)
        bot.register_next_step_handler(opcao, editarFunc4)
    except ValueError:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/editar')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,
                         "Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n\n Tente novamente com \n-> /editar \n\n Caso queira voltar para o início, toque em \n-> /comecar",
                         reply_markup=voltar)
    except KeyError:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/editar')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,
                         "Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n\n Tente novamente com \n-> /editar \n\n Caso queira voltar para o início, toque em \n-> /comecar",
                         reply_markup=voltar)

def editarFunc4(message):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    chat_id = message.chat.id
    wb = load_workbook(arquivo[0])
    ws = wb[mesSelecionado[0]]
    if message.text == "Nome":
        opcaoNumero.append("nome")
        markup = types.ForceReply()
        a = get_column_letter(1)
        cell = [a + valorNomeEditar[0]]
        valorNomeEditar2.append(cell)
        opcao = bot.send_message(chat_id, "Digite o novo nome para o gasto", reply_markup=markup)
        bot.register_next_step_handler(opcao, editarFunc5)
    elif message.text == "Valor":
        opcaoNumero.append("confirma")
        markup = types.ForceReply()
        b = get_column_letter(2)
        cell = [b + valorNomeEditar[0]]
        valorNomeEditar2.append(cell)
        opcao = bot.send_message(chat_id, "Digite o novo valor para o gasto", reply_markup=markup)
        bot.register_next_step_handler(opcao, editarFunc5)
    else:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/editar')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,"Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n\n Tente novamente com \n-> /editar \n\n Caso queira voltar para o início, toque em \n-> /comecar", reply_markup=voltar)
def editarFunc5(message):
    from openpyxl import load_workbook
    chat_id = message.chat.id
    pessoa = message.from_user.first_name
    wb = load_workbook(arquivo[0])
    ws = wb[mesSelecionado[0]]
    mensagem1 = message.text
    listaMensagem = list(mensagem1)
    i = 0
    if opcaoNumero[0] != 'confirma':
        while i < len(listaMensagem):
            if listaMensagem[i] == ",":
                listaMensagem[i] = "."
            i += 1
        mensagem1 = ''.join(listaMensagem)
        ws[valorNomeEditar2[0][0]].value = str(mensagem1)
        mesSelecionado.clear()
        valorNomeEditar.clear()
        valorNomeEditar2.clear()
        opcaoNumero.clear()
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/editar')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,"Para editar mais alguma coisa, aperte \n-> /editar \n\n Para voltar ao início, aperte \n-> /comecar",reply_markup=voltar)
        wb.save(f"gastos de {pessoa} {arquivo[0]}.xlsx")
    else:
        from integracaoExcel import floatcheck
        if floatcheck(message.text):
            while i < len(listaMensagem):
                if listaMensagem[i] == ",":
                    listaMensagem[i] = "."
                i += 1
            mensagem1 = ''.join(listaMensagem)
            ws[valorNomeEditar2[0][0]].value = str(mensagem1)
            mesSelecionado.clear()
            valorNomeEditar.clear()
            valorNomeEditar2.clear()
            opcaoNumero.clear()
            voltar = types.ReplyKeyboardMarkup()
            a = types.KeyboardButton('/editar')
            b = types.KeyboardButton('/comecar')
            voltar.row(a, b)
            bot.send_message(chat_id,
                             "Para editar mais alguma coisa, aperte \n-> /editar \n\n Para voltar ao início, aperte \n-> /comecar", reply_markup=voltar)
            wb.save(f"gastos de {pessoa} {arquivo[0]}.xlsx")
        else:
            mesSelecionado.clear()
            valorNomeEditar.clear()
            valorNomeEditar2.clear()
            opcaoNumero.clear()
            chat_id = message.chat.id
            voltar = types.ReplyKeyboardMarkup()
            a = types.KeyboardButton('/editar')
            b = types.KeyboardButton('/comecar')
            voltar.row(a, b)
            bot.send_message(chat_id,
                             "Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n\n Tente novamente com \n-> /editar \n\n Caso queira voltar para o início, toque em \n-> /comecar",
                             reply_markup=voltar)


@bot.message_handler(commands=["contasMes"])
def totalDoMes1(message):
    chat_id = message.chat.id
    pessoa = message.from_user.first_name
    meses = types.ReplyKeyboardMarkup()
    a = types.KeyboardButton('jan'.capitalize())
    b = types.KeyboardButton('fev'.capitalize())
    c = types.KeyboardButton('mar'.capitalize())
    d = types.KeyboardButton('abr'.capitalize())
    e = types.KeyboardButton('mai'.capitalize())
    f = types.KeyboardButton('jun'.capitalize())
    g = types.KeyboardButton('jul'.capitalize())
    h = types.KeyboardButton('ago'.capitalize())
    i = types.KeyboardButton('set'.capitalize())
    j = types.KeyboardButton('out'.capitalize())
    k = types.KeyboardButton('nov'.capitalize())
    l = types.KeyboardButton('dez'.capitalize())
    m = types.KeyboardButton(seuMes().capitalize())
    meses.row(m)
    meses.row(a, b, c, d)
    meses.row(e, f, g, h)
    meses.row(i, j, k, l)
    mes = bot.send_message(chat_id,"Selecione o mês \n\n (obs: a primeira e maior opção será o mês atual)", reply_markup=meses)
    bot.register_next_step_handler(mes, totalDoMes2)

def totalDoMes2(message):
    try:
        chat_id = message.chat.id
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import datetime
        pessoa = message.from_user.first_name
        data = str(datetime.datetime.now())
        wb = load_workbook(arquivo[0])
        ws = wb[message.text]
        b = 0
        coluna = get_column_letter(2)
        for row in range(1, 50):
            valor = ws[coluna + str(row)].value
            if valor == None:
                valor = 0.0
            b += float(valor)
        bot.send_message(chat_id, "O total gasto em {message.text} foi de R${b}")
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/contasMes')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id, "Para ver o total de outro mês, toque em \n-> /contasMes \n\n Para voltar ao início, toque em \n-> /comecar", reply_markup=voltar)

        wb.save(f"gastos de {pessoa} {arquivo[0]}.xlsx")
    except KeyError:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/contasMes')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,"Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n Tente novamente com \n-> /contasMes \n\n Caso queira voltar para o início, toque em \n-> /comecar",reply_markup=voltar)

@bot.message_handler(commands=["addGasto"])
def selecionarMes(message):
    markup = types.ForceReply()
    chat_id = message.chat.id
    valorNome = bot.send_message(chat_id, "Digite o nome do gasto: ", reply_markup=markup)
    bot.register_next_step_handler(valorNome, editorExcel)


def editorExcel(message):
    markup = types.ForceReply()
    valorExcel.append(message.text)
    chat_id = message.chat.id
    valorGasto = bot.send_message(chat_id, "Digite o valor gasto: ", reply_markup=markup)
    bot.register_next_step_handler(valorGasto, seletorMes)


def seletorMes(message):
    i = 0
    mensagem = message.text
    listaMensagem = list(mensagem)
    from integracaoExcel import floatcheck
    if floatcheck(message.text):
        while i < len(listaMensagem):
            if listaMensagem[i] == ",":
                listaMensagem[i] = "."

            i += 1
        mensagem = ''.join(listaMensagem)
        valorExcel.append(mensagem)
        chat_id = message.chat.id
        meses = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('Jan'.capitalize())
        b = types.KeyboardButton('fev'.capitalize())
        c = types.KeyboardButton('mar'.capitalize())
        d = types.KeyboardButton('abr'.capitalize())
        e = types.KeyboardButton('mai'.capitalize())
        f = types.KeyboardButton('jun'.capitalize())
        g = types.KeyboardButton('jul'.capitalize())
        h = types.KeyboardButton('ago'.capitalize())
        i = types.KeyboardButton('set'.capitalize())
        j = types.KeyboardButton('out'.capitalize())
        k = types.KeyboardButton('nov'.capitalize())
        l = types.KeyboardButton('dez'.capitalize())
        m = types.KeyboardButton(seuMes().capitalize())
        meses.row(m)
        meses.row(a, b, c, d)
        meses.row(e, f, g, h)
        meses.row(i, j, k, l)
        meses = bot.send_message(chat_id,"Selecione o mês \n\n (obs: a primeira e maior opção será o mês atual)", reply_markup=meses)
        bot.register_next_step_handler(meses, final)
    else:
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/addGasto')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,"Algo deu errado. Por favor, confirme o valor do gasto e tente novamente com \n-> /addGasto \n\n Caso queira voltar para o início, toque em \n-> /comecar", reply_markup=voltar)

def final(message):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from integracaoExcel import floatcheck
        import datetime
        chat_id = message.chat.id
        wb = load_workbook(arquivo[0])
        ws = wb[message.text]
        ws.append(valorExcel)
        coluna1 = get_column_letter(1)
        coluna2 = get_column_letter(2)
        print(valorExcel)
        print(ws[coluna1 + "51"].value, ws[coluna2 + "51"].value)
        if ws[coluna1 + "51"].value is None and ws[coluna2 + "51"].value is None and floatcheck(valorExcel[1]):
            pessoa = message.from_user.first_name
            voltar = types.ReplyKeyboardMarkup()
            a = types.KeyboardButton('/comecar')
            voltar.row(a)
            wb.save(arquivo[0])
            print(valorExcel)
            valorExcel.clear()
            print(valorExcel)
            bot.send_message(chat_id,"Gasto adicionado! \n\n Para adicionar outro gasto, toque em \n-> /addGasto \n\n Para iniciar novamente, toque em \n-> /comecar",reply_markup=voltar)
        else:
            valorExcel.clear()
            voltar = types.ReplyKeyboardMarkup()
            a1 = types.KeyboardButton('/editarVer')
            b2 = types.KeyboardButton('/addGasto')
            c3 = types.KeyboardButton('/comeco')
            voltar.row(a1,b2,c3)
            bot.send_message(chat_id,"Aparentemente, você estourou o limite de gastos (50) do mês {message.text} ou digitou algo errado. \n\n Para editar os gastos deste mês, toque em \n-> /editarVer e selecione o mês \n\n Para tentar adicionar novos gastos em outros meses, toque em \n-> /addGasto e selecione outro mês \n\n Para voltar ao início, toque em \n-> /comecar", reply_markup=voltar)
    except KeyError:
        valorExcel.clear()
        chat_id = message.chat.id
        voltar = types.ReplyKeyboardMarkup()
        a = types.KeyboardButton('/addGasto')
        b = types.KeyboardButton('/comecar')
        voltar.row(a, b)
        bot.send_message(chat_id,"Algo deu errado. Por favor, leia atentamente as mensagens que o bot enviar. \n Tente novamente com \n-> /addGasto \n\n Caso queira voltar para o início, toque em \n-> /comecar",reply_markup=voltar)

@bot.message_handler(func=lambda m: True)
def echo_all(message):
    chat_id = message.chat.id
    voltar = types.ReplyKeyboardMarkup()
    a = types.KeyboardButton('/comecar')
    voltar.row(a)
    bot.send_message(chat_id, "Para iniciar, toque em /comecar", reply_markup=voltar)


bot.infinity_polling()
