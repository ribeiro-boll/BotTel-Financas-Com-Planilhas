

class Gasto:
    def __init__(self, nome):
        self.nome = nome

    def adicionar_valor(self, valor):
        self.valor = valor

    def retornar_valor_gasto(self):
        return self.valor

    def gerar_lista(self):
        valor_cell = (self.nome, self.valor)
        return valor_cell

    def adicionar_gasto(self, arquivo, message, lista, valor):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from integracaoExcel import floatcheck
        wb = load_workbook(arquivo)
        ws = wb[message.text]
        ws.append(lista)
        coluna1 = get_column_letter(1)
        coluna2 = get_column_letter(2)
        if ws[coluna1 + "51"].value is None and ws[coluna2 + "51"].value is None and floatcheck(valor):
            wb.save(arquivo)
            return "1"
        else:
            return "0"
class Planilha:
    def __init__(self, nome):
        self.nome = nome

    def definir_ano_inicial(self):
        import datetime
        ano = datetime.datetime.now().year
        self.arquivo = f"gastos de {self.nome} {ano}.xlsx"

    def mudar_ano(self, ano):
        self.arquivo = f"gastos de {self.nome} {ano}.xlsx"

    def retornar_nome_arquivo_planilha_atual(self):
        return self.arquivo

def listar_gastos(arquivo, message):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(arquivo)
    ws = wb[message.text]
    coluna = get_column_letter(2)
    colunaNome = get_column_letter(1)
    colunaTexto = ''
    for row in range(1, 50):
        valor = ws[colunaNome + str(row)].value
        if valor == None:
            continue
        else:
            colunaTexto += f'{ws[colunaNome + str(row)].value} = R${float(ws[coluna + str(row)].value)}\n\n'
    return colunaTexto
def somatorio(arquivo, message):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(arquivo)
    ws = wb[message.text]
    b = 0.0
    coluna = get_column_letter(2)
    for row in range(1, 50):
        valor = ws[coluna + str(row)].value
        if valor == None:
            valor = 0.0
        b += float(valor)
    return b


def funcao_opcao_mes(id, bot):
    from telebot import types
    meses = types.ReplyKeyboardMarkup()
    a = types.KeyboardButton('jan'.capitalize())
    b = types.KeyboardButton('fev'.capitalize())
    c = types.KeyboardButton('mar'.capitalize())
    d = types.KeyboardButton('abr'.capitalize())
    e = types.KeyboardButton('mai'.capitalize())
    f = types.KeyboardButton('jun'.capitalize())
    g = types.KeyboardButton('jul'.capitalize())
    h = types.KeyboardButton('ago'.capitalize())
    i = types.KeyboardButton('set'.capitalize())
    j = types.KeyboardButton('out'.capitalize())
    k = types.KeyboardButton('nov'.capitalize())
    l = types.KeyboardButton('dez'.capitalize())
    m = types.KeyboardButton(seuMes().capitalize())
    meses.row(m)
    meses.row(a, b, c, d)
    meses.row(e, f, g, h)
    meses.row(i, j, k, l)
    mes = bot.send_message(id,"Selecione o mês \n\n (obs: a primeira e maior opção será o mês atual)", reply_markup=meses)
    return mes

def seuMes():
    import datetime
    a = str(datetime.date.today())[5] + str(datetime.date.today())[6]
    match int(a):
        case(1):
            return 'jan'
        case(2):
            return 'fev'
        case (3):
            return "mar"
        case (4):
            return "abr"
        case (5):
            return "mai"
        case (6):
            return "jun"
        case (7):
            return "jul"
        case (8):
            return "ago"
        case(9):
            return "set"
        case (10):
            return "out"
        case (11):
            return "nov"
        case (12):
            return "des"

def user(message):
    chat_id = message.chat.id
    return chat_id
def nome_pessoa(message):
    pessoa = message.from_user.first_name
    return pessoa

def escolher_ano():
    import datetime
    ano_base = datetime.datetime.now().year
    ano_mais = int(ano_base)
    ano_menos = int(ano_base) - 4
    lista_anos = []
    lista_anos.clear()
    for i in range(3):
        ano_menos += 1
        lista_anos.append(ano_menos)
    lista_anos.append(ano_base)
    for j in range(3):
        ano_mais += 1
        lista_anos.append(ano_mais)
    return lista_anos

def criarArquivoAno(arquivo):
    from openpyxl import Workbook
    try:
        open(arquivo, 'r')
        return arquivo
    except IOError:
        resumoAnual = Workbook()
        meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
        for mes in meses:
            resumoAnual.create_sheet(mes.capitalize())
        resumoAnual.save(arquivo)
        return arquivo

def listar_edicao(message, arquivo):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    a = get_column_letter(1)
    wb = load_workbook(arquivo)
    ws = wb[message]
    nomes = ''
    for row in range(1, 50):
        valor = ws[a + str(row)].value
        if valor == None:
            continue
        else:
            nomes += f'{ws[a + str(row)].value} gasto nº /0000{row}\n\n'
    return nomes

def funcao_editar_selecionar(message, arquivo, mes):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    var = list(message.text)
    var.remove("/")
    var.remove("0")
    var.remove("0")
    var.remove("0")
    var.remove("0")
    mesagem = str(''.join(var))
    wb = load_workbook(arquivo)
    ws = wb[mes]
    a = get_column_letter(1)
    nnome = ws[a + mesagem].value
    lista_resposta = [nnome, mesagem]
    return lista_resposta

def floatcheck(a):
    b = list(str(a))
    c = 0
    for i in b:
        if i in "0123456789.,":
            c += 1
    if c == len(a):
        return True
    else:
        return False
